import json

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side

from simple_loggers import SimpleLogger

from pypubmed.util import safe_open


class Export(object):

    logger = SimpleLogger('Export')

    def __init__(self, data, outfile='out.xlsx', outtype=None, fields=None, **kwargs):
        self.outfile = outfile
        self.outtype = outtype or outfile.split('.')[-1]
        self.data = list(self.filter_data(data, fields)) if fields else data

    def filter_data(self, data, fields):
        fields = fields.strip(',').split(',')
        for context in data:
            out_ctx = {k: v for k, v in context.items() if k in fields}
            yield out_ctx

    def export(self):
        if self.outtype == 'xlsx':
            self.export_xlsx()
        elif self.outtype == 'json':
            self.export_json()
        elif self.outtype in ('jl', 'jsonlines'):
            self.export_json_lines()
        else:
            self.logger.error('outtype is invalid, please check!')
            exit(1)
        
        self.logger.info('save file: {}'.format(self.outfile))

    def write_title(self, sheet, titles, fg_color=colors.BLACK, bg_color=colors.WHITE, border=True, bold=True, width=18, size=12):
        for col, value in enumerate(titles, 1):

                w = width * 4 if value in ('abstract', 'abstract_cn') else width

                sheet.column_dimensions[get_column_letter(col)].width = w
                _ = sheet.cell(1, col, value=value)

                # set cell styles
                _.alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
                _.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                _.font = Font(bold=bold, color=fg_color, size=size)

    def add_hyperlink(self, key, value):
        """
            method1:
                value=HYPERLINK("http://www.baidu.com", "baidu")
                sheet.cell(row, colum, value=value)

            method2:
                _ = sheet.cell(row, column, value='baidu')
                _.hyperlink = 'http://www.baidu.com'
        """
        if key == 'pmid':
            url = 'https://pubmed.ncbi.nlm.nih.gov/{value}/'
        elif key == 'pmc':
            url = 'https://www.ncbi.nlm.nih.gov/pmc/articles/{value}/'
        elif key == 'doi':
            url = 'https://doi.org/{value}'
        else:
            return None

        url = url.format(**locals())

        return url
    
    @property
    def all_fields(self):
        return '''
            pmid title abstract abstract_cn impact_factor journal med_abbr iso_abbr pubdate pubmed_pubdate
            pmc issn e_issn doi year pagination volume issue   
            pub_status authors keywords pub_types cited
        '''.split()

    def sort(self, item):
        if type(item) == tuple:
            k = item[0]
        else:
            k = item

        if k in self.all_fields:
            return self.all_fields.index(k)
        else:
            return 9999

    def export_xlsx(self, sheet_title='Result', freeze_panes='B1'):
        """
            PatternFill:
            - https://openpyxl.readthedocs.io/en/latest/api/openpyxl.styles.fills.html?highlight=PatternFill
        """
        book = openpyxl.Workbook()
        sheet = book.active

        # # freeze the first column
        # sheet.freeze_panes = 'B1'

        # # freeze the first row
        # sheet.freeze_panes = 'A2'

        # freeze the first column and the first row
        sheet.freeze_panes = 'B2'

        sheet.title = sheet_title

        titles = sorted(list(self.data[0].keys()), key=self.sort)
        self.write_title(sheet, titles)
        
        for row, context in enumerate(self.data, 2):

            color = '00b3ffb3' if row % 2 else '00b3ffff'
            for col, (key, value) in enumerate(sorted(list(context.items()), key=self.sort), 1):
                if type(value) == list:
                    try:
                        value = ', '.join(value)
                    except:
                        try:
                            value = json.dumps(value, ensure_ascii=False)
                        except:
                            value = str(value)
                elif type(value) == dict:
                    value = json.dumps(value, ensure_ascii=False)
                    
                _ = sheet.cell(row, col, value=value)
                _.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                if key in ('pmid', 'pmc', 'doi') and value != '.':
                    _.hyperlink = self.add_hyperlink(key, value)
                    _.font = Font(color=colors.BLUE, italic=True)

                wrap_text = None
                if key in ('abstract', 'abstract_cn'):
                    wrap_text = True
                
                _.alignment = Alignment(horizontal='left',vertical='center',wrap_text=wrap_text)

        book.save(self.outfile)

    def export_json(self, **kwargs):
        with safe_open(self.outfile, 'w') as out:
            json.dump(self.data, out, ensure_ascii=False, **kwargs)

    def export_json_lines(self):
        with safe_open(self.outfile, 'w') as out:
            for context in self.data:
                out.write(json.dumps(context, ensure_ascii=False) + '\n')


if __name__ == '__main__':
    data = [
        {'pmid':1, 'pmc': 'PMC4472309', 'doi': '10.1016/0006-2952(75)90018-0'},
        {'pmid':1, 'pmc': '.', 'doi': '10.1016/0006-2952(75)90018-0'},
        {'pmid':1, 'pmc': 'PMC4472309', 'doi': '.'},
        {'pmid':1, 'pmc': 'PMC4472309', 'doi': '10.1016/0006-2952(75)90018-0'},
        {'pmid':1, 'pmc': 'PMC4472309', 'doi': '10.1016/0006-2952(75)90018-0'},
    ]

    Export(data).export()
